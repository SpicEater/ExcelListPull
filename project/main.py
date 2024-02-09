import re
import json
import openpyxl
import os.path
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


class Excel(object):
    obs = 'D:\\ВВ\\Егор\\Нужные файлы\\СинХРОН\\История просмотров 2.md'
    ex = 'D:\\ВВ\\Егор\\Нужные файлы\\Список просмотров.xlsx'
    save_ex = ex
    book = openpyxl.load_workbook(filename=ex)
    list = book.active

    def __init__(self, obs = obs, ex = ex, save_ex = save_ex):
        self.ex = ex
        self.obs = obs
        self.save_ex = save_ex

    def export_obs(self, text=obs):

        symbols = []

        with open(text, 'r', encoding="utf-8") as file:
            for sym in file:
                translat = sym.strip('\n \t')
                if translat[:2] == 'Н:' or translat[:2] == 'К:':
                    symbols.append(re.sub(r'\s+', '', sym))
                elif translat[:4] == 'НиК:':
                    reg = translat.lstrip('НиК: ')
                    symbols.append('Н:' + reg)
                    symbols.append('К:' + reg)
                else:
                    symbols.append(translat)

        return symbols

    def time_format(seld, date_time):

        date = f'{date_time[:-2]}'

        if date != '':
            day, mon, year = date_time.split('.')
            if mon[1:] == "":
                mon = "0" + mon
            date = f'{day}.{mon}.20{year}'

        return date

    def max_counter(self, name):

        try:
            name_new = name.values()
            counter = max(name_new)
        except:
            name = {}
            counter = 1

        while True:
            counter += 1
            ferst_num = self.list[f'B{counter}'].value
            if ferst_num == None:
                counter -= 1
                break
            name[ferst_num] = counter

        with open('saved_dictionary.json', 'w') as f:
            json.dump(name, f, indent=4)

        return counter + 1

    def int_time(self, id, obc, ob):

        thins = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='thin'),
                       bottom=Side(style='thin'))
        d = self.list[f"D{id}"]
        e = self.list[f"E{id}"]
        d.alignment = Alignment(horizontal='right', vertical="center")
        e.alignment = Alignment(horizontal='right', vertical="center")
        d.border = thins
        e.border = thins

        self.list[f"D{id}"] = self.time_format(obc[ob + 1][2:])
        self.list[f"E{id}"] = self.time_format(obc[ob + 2][2:])

    def start(self):

        name_dic = {}
        self.max_counter(name_dic)

        if os.path.isfile('saved_dictionary.json'):
            with open('saved_dictionary.json', 'r') as f:
                name_dic = json.load(f)
        for ob in range(0, len(self.export_obs()), 3):
            name_obc = self.export_obs()
            max_count = self.max_counter(name_dic)
            if name_obc[ob] in name_dic:
                id = name_dic[name_obc[ob]]
                self.int_time(id, name_obc, ob)
            else:
                thins = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='thin'),
                               bottom=Side(style='thin'))
                alignment = Alignment(horizontal='center', vertical="center", wrap_text=True)
                fill = PatternFill('solid', fgColor='d9d9d9')

                a = self.list[f"A{max_count}"]

                a.alignment = alignment
                a.fill = fill
                a.border = thins
                a.font = Font(bold=True)

                self.list[f"B{max_count}"].border = thins
                self.list[f"B{max_count}"].alignment = alignment

                self.list[f"A{max_count}"] = max_count - 1
                self.list[f"B{max_count}"] = name_obc[ob]
                self.int_time(max_count, name_obc, ob)
                max_count += 1
        self.book.save(self.save_ex)
        print('badanCadan')


if __name__ == "__main__":
    s = Excel()
    s.start()